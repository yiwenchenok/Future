import os.path

import requests
from bs4 import BeautifulSoup
import json
import random
import time
from openpyxl import Workbook, load_workbook

# 建立一个user-Agent池防屏蔽
user_agents = [
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60',
    'Opera/8.0 (Windows NT 5.1; U; en)',
    'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0',
    'Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2 ',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)',
    'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0',
    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0) ']


def get_ws(file_name):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        title = ['product_id', 'p_name', 'p_price', 'p_imag_scr', 'p_shop', 'p_comment_count', 'p_pinpai', 'p_detail']
        wb = Workbook()
        ws = wb.active
        ws.append(title)
    return wb, ws


def write_context(ws, data):
    # 数据
    ws.append(data)


def run(search_name='飞机杯', max_page=100, max_count=100, file_name='result.xlsx', cookie=''):
    """
    爬取信息
    :param search_name: 需要爬取的商品名
    :param max_page: 获取最大的页码数
    :param max_count: 获取多少商品数量
    :param file_name: 保存结果的文件名
    :param cookie: 如果不能获取需要添加cookie
    :return:
    """
    # 随机生成一个headers
    print('start')
    user_agent = random.choice(user_agents)
    headers = {"User-Agent": user_agent, "Referer": "https://www.jd.com/"}
    if cookie:
        headers['cookie'] = cookie
    wb, ws = get_ws(file_name)
    index = 0
    for page in range(1, max_page):
        url = 'https://search.jd.com/Search?keyword={}&suggest=2.his.0.0&wq={}&page={}&s=1&click=0'.format(search_name,
                                                                                                           search_name,
                                                                                                           page)
        time.sleep(random.randint(1, 10))
        html = requests.get(url, headers=headers).content
        soup = BeautifulSoup(html, 'html.parser')
        soup_allli = soup.find_all('li', class_="gl-item")
        for soup_li in soup_allli:
            product_id = soup_li['data-sku']
            p_name = soup_li.find('div', class_="p-name").text.strip()
            p_price = soup_li.find('div', class_="p-price").strong.i.text
            p_img_obj = soup_li.find('div', class_='p-img')
            p_imag_scr = 'https:' + p_img_obj.a.img['data-lazy-img']
            p_shop = soup_li.find('div', class_='p-shop').text.strip()
            p_detail_src = 'https:' + p_img_obj.a['href']
            p_commit_scr = 'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId={}&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&fold=1'.format(
                product_id)
            # 获取评价个数
            time.sleep(random.randint(1, 20))
            comment_html = requests.get(p_commit_scr, headers=headers).text
            res_data = json.loads(comment_html[20:-2])
            p_comment_count = res_data['productCommentSummary']['commentCountStr']
            p_pinpai, p_detail = get_shangpin_detail(product_id)
            data = [product_id, p_name, p_price, p_imag_scr, p_shop, p_comment_count, p_pinpai, json.dumps(p_detail)]
            print(p_name, p_price, p_detail_src, p_imag_scr, p_shop, p_comment_count, sep='\n')
            ws.append(data)
            index += 1
            if index > max_count:
                wb.save(file_name)
                print('done ')
                return ''
        wb.save(file_name)
    wb.save(file_name)


def get_shangpin_detail(product_id):
    """
    获取商品详情
    :param product_id: 商品id
    :return: 品牌名，商品详情
    """
    time.sleep(random.randint(3, 10))
    user_agent = random.choice(user_agents)
    headers = {"User-Agent": user_agent, "Referer": "https://www.jd.com/"}
    url = f'https://item.jd.com/{product_id}.html'
    html = requests.get(url, headers=headers).content
    soup = BeautifulSoup(html, 'html.parser')
    p_pinpai = soup.select('#parameter-brand li')[0]['title']  # 品牌
    p_detail = [li.text for li in soup.select('.parameter2 li')]
    p_detail_dict = []
    for i in p_detail:
        key, val = i.split('：')
        p_detail_dict.append({key: val})
    return p_pinpai, p_detail_dict


if __name__ == '__main__':
    run(search_name='飞机杯', max_count=5, file_name='result.xlsx')
