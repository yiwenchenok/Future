import os.path

import requests
from bs4 import BeautifulSoup
import json
from urllib import parse
import random
import time
from openpyxl import Workbook, load_workbook

# 建立一个user-Agent池防屏蔽
user_agents = [
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60',
    'Opera/8.0 (Windows NT 5.1; U; en)',
    'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0',
    'Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2 ',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)',
    'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0',
    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0) ']

headers = {
    'Connection': 'keep-alive',
    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="96", "Google Chrome";v="96"',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'X-Requested-With': 'XMLHttpRequest',
    'sec-ch-ua-mobile': '?0',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
    'sec-ch-ua-platform': '"macOS"',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Dest': 'empty',
    'Referer': '',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}


def get_ws(file_name):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        title = ['coid', 'company_name', 'job_name', 'workarea_text', 'providesalary_text', 'issuedate',
                 'companytype_text',
                 'companysize_text', 'companyind_text', 'job_href', 'jobwelf', 'attribute_text']
        wb = Workbook()
        ws = wb.active
        ws.append(title)
    return wb, ws


def write_context(ws, data):
    # 数据
    ws.append(data)


def run(search_job='+', start_page=1, max_page=100, max_count=100, file_name='job_result.xlsx', cookie=None):
    """
    爬取信息
    :param search_job: 职位名
    :param start_page: 起始页
    :param max_page: 获取最大的页码数
    :param max_count: 获取多少商品数量
    :param file_name: 保存结果的文件名
    :param cookie: 如果不能获取需要添加cookie
    :return:
    """
    error_url_list = []
    # 随机生成一个headers
    search_job = parse.quote(search_job)
    base_url = 'https://search.51job.com/list/000000,000000,0000,00,9,99,{search_job},2,{page}.html?lang=c&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare='
    print('start')
    user_agent = random.choice(user_agents)
    headers.update({"User-Agent": user_agent})
    session = requests.Session()
    if cookie:
        headers['Cookie'] = cookie
    wb, ws = get_ws(file_name)
    index = 0
    for page in range(start_page, max_page + 1):
        print(f'开始获取第{page}页内容')
        url = base_url.format(search_job=search_job, page=page)
        headers.update({'Referer': url})
        time.sleep(random.randint(5, 10))
        # 5. 分析结果 公司名、岗位名、工作地址、薪资、发布时间、工作描述、公司类型、员工人数、所属行业
        try:
            res = session.get(url=url, headers=headers)
            res_json = eval(res.text)
            engine_jds = res_json.get('engine_jds')
            for job_info in engine_jds:
                # 6、 字段解析
                coid = job_info.get('coid')  # 岗位id
                company_name = job_info.get('company_name')  # 公司名字
                job_name = job_info.get('job_name')  # 岗位名字
                workarea_text = job_info.get('workarea_text')  # 工作地址
                providesalary_text = job_info.get('providesalary_text').replace("\\", '')  # 薪资
                issuedate = job_info.get('issuedate')  # 发布时间
                companytype_text = job_info.get('companytype_text').replace('\\', '')  # 公司类型
                companysize_text = job_info.get('companysize_text')  # 员工人数
                companyind_text = job_info.get('companyind_text').replace('\\', '')  # 所属行业
                job_href = job_info.get('job_href').replace('\\', '')  # 岗位信息
                jobwelf = job_info.get('jobwelf')  # 公司福利
                attribute_text = '|'.join(job_info.get('attribute_text'))  # 岗位要求
                # 7、保存信息到Excel中
                ws.append(
                    [coid, company_name, job_name, workarea_text, providesalary_text, issuedate, companytype_text,
                     companysize_text, companyind_text, job_href, jobwelf, attribute_text])
                index += 1
                if index > max_count:
                    wb.save(file_name)
                    print('done ')
                    return ''
            wb.save(file_name)
        except Exception as e:
            print(e)
            print(f'cookie 失效, 获取{page}页失败,请设置start_page={page}和cookie, 重新爬取')
            break


if __name__ == '__main__':
    cookie = '_uab_collina=165184878215205617976495; guid=258301ecc190aa20c07e60f3103136b2; nsearch=jobarea%3D%26%7C%26ord_field%3D%26%7C%26recentSearch0%3D%26%7C%26recentSearch1%3D%26%7C%26recentSearch2%3D%26%7C%26recentSearch3%3D%26%7C%26recentSearch4%3D%26%7C%26collapse_expansion%3D; search=jobarea%7E%60000000%7C%21ord_field%7E%600%7C%21recentSearch0%7E%60000000%A1%FB%A1%FA000000%A1%FB%A1%FA0000%A1%FB%A1%FA00%A1%FB%A1%FA99%A1%FB%A1%FA%A1%FB%A1%FA99%A1%FB%A1%FA99%A1%FB%A1%FA99%A1%FB%A1%FA99%A1%FB%A1%FA9%A1%FB%A1%FA99%A1%FB%A1%FA%A1%FB%A1%FA0%A1%FB%A1%FA%7Bsearch_job%7D%A1%FB%A1%FA2%A1%FB%A1%FA1%7C%21; _ujz=MTkyNjE0MTIzMA%3D%3D; slife=lowbrowser%3Dnot%26%7C%26lastlogindate%3D20220506%26%7C%26; ps=needv%3D0; 51job=cuid%3D192614123%26%7C%26cusername%3DXbHr3rM6s3aljy5maUrJ7I8d2eeMfHCgOBuvsVkXZGs%253D%26%7C%26cpassword%3D%26%7C%26cname%3DvScdfZN%252B%252BS6DWGKe4Bi1LA%253D%253D%26%7C%26cemail%3DEeIFGEsj4AG6msAnR0YKorVe6zyc%252BOmAYxkuGsSgarI%253D%26%7C%26cemailstatus%3D0%26%7C%26cnickname%3D%26%7C%26ccry%3D.0JMFYV.tme2Y%26%7C%26cconfirmkey%3D18GxbVwRCiwpE%26%7C%26cautologin%3D1%26%7C%26cenglish%3D0%26%7C%26sex%3D0%26%7C%26cnamekey%3D181BOuQbFOCKY%26%7C%26to%3D8a6ae130e5d18dbbac4fdac6be2ec2766275356a%26%7C%26; acw_tc=76b20ff416518487819155848e5ea1c3fd9f3776ee2ab53cf408899bd69cd0; acw_sc__v2=6275364d5c90b029576a01d45c0280c42533c430; ssxmod_itna=eqmxuDnDgDciv4BP8e5Dk7qDtA=BI6xPD7YI9jDBkGo4iNDnD8x7YDv+IZYwYQOi4mfGKjAB5dhm7hmE+d3W2jHPDHxY=DUOBWhYD4SKGwD0eG+DD4DWUx03DoxGYBGx0bSyuLvzQGRD0YDzqDgD7QVbqDEDG3D03=qYg4bbTtjYnNqxGtqmkGxbqDMD7tD/+xTLeDBt=ab0TtNxPGWWOQxPeGuDG6ehTdqx0P9nwt0Ac5eWhDe+A54bhxFQGqPiGeqU0OqW7DN/B+=YE+QK0Dq/GhNmA+fzKDAi0pzQiD==; ssxmod_itna2=eqmxuDnDgDciv4BP8e5Dk7qDtA=BI6xPD7YI9D8d17eGXxeIdGaCWAkaCrAoErBRGN8K9Bf9N=UDnRD+sznlgX1tozlyaKToRZSXd=oiWlcLIo5zQ6LsSfnp9vIEbTllCmW7izMvx1KOBc9iccKOXoB112nG/hKCniD438yDmYgrGpqHVleqZOv7w4TtQfKwpI3EiOS8N2WN63df6aqhQhLdm2nKQfcU8=k3DQFIDjKD+pEoNi8QKYD='
    run(search_job='Java', max_count=50, cookie=cookie)
